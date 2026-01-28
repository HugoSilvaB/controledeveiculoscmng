import os
import io
from datetime import datetime
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from wtforms import StringField, FloatField, SubmitField, SelectField, PasswordField
from wtforms.validators import DataRequired, Length, Optional
from flask_wtf.file import FileField, FileRequired, FileAllowed
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
from PIL import Image
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import current_app, send_file, abort



# --- CONFIGURAÇÃO INICIAL ---
load_dotenv()
app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY') or 'camara_secret_123'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# --- BANCO DE DADOS E LOGIN ---
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Usuario, int(user_id))

LISTA_GABINETES = [
    ('Presidência - Charles do Oceano', 'Presidência - Charles do Oceano'),
    ('Gabinete - André Logos', 'Gabinete - André Logos'),
    ('Gabinete - Chico Civil', 'Gabinete - Chico Civil'),
    ('Gabinete - Cleiton Lucio', 'Gabinete - Cleiton Lucio'),
    ('Gabinete - Emilio Santiago', 'Gabinete - Emilio Santiago'),
    ('Gabinete - Ewerton Vidal', 'Gabinete - Ewerton Vidal'),
    ('Gabinete - Fabricio Chaves', 'Gabinete - Fabricio Chaves'),
    ('Gabinete - Luana Marques', 'Gabinete - Luana Marques'),
    ('Gabinete - Marcos Oliveira', 'Gabinete - Marcos Oliveira'),
    ('Gabinete - Marquim do Baxim', 'Gabinete - Marquim do Baxim'),
    ('Gabinete - Pacífico', 'Gabinete - Pacífico'),
    ('Gabinete - Paulo Jordão', 'Gabinete - Paulo Jordão'),
    ('Gabinete - Renato Caldas', 'Gabinete - Renato Caldas'),
    ('Gabinete - Willian Faleiro', 'Gabinete - Willian Faleiro'),
    ('Gabinete - Zé Lopes', 'Gabinete - Zé Lopes'),
]

# --- MODELOS ---
class Usuario(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100))
    cpf = db.Column(db.String(11), unique=True, nullable=False)
    senha = db.Column(db.String(200), nullable=False)
    cargo = db.Column(db.String(50))
    ativo = db.Column(db.Boolean, default=True)
    gabinete = db.Column(db.String(100))

class Veiculo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    modelo = db.Column(db.String(50), nullable=False)
    placa = db.Column(db.String(10), default="OFICIAL")
    km_atual = db.Column(db.Integer, default=0) # Esta é a coluna que faltava
    km_revisao_proxima = db.Column(db.Integer, default=10000) # E esta também

class RegistroUso(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    gabinete_vereador = db.Column(db.String(100), nullable=False)
    motorista_nome = db.Column(db.String(100), nullable=False)
    veiculo_id = db.Column(db.Integer, db.ForeignKey('veiculo.id'), nullable=False)
    data_hora_saida = db.Column(db.DateTime, nullable=False, default=datetime.now)
    km_saida = db.Column(db.Float, nullable=False)
    foto_km_saida = db.Column(db.String(255), nullable=False)
    data_hora_chegada = db.Column(db.DateTime, nullable=True)
    km_chegada = db.Column(db.Float, nullable=True)
    foto_km_chegada = db.Column(db.String(255), nullable=True)
    destino_finalidade = db.Column(db.String(255), nullable=True)
    observacoes = db.Column(db.String(500), nullable=True)
    foto_ocorrencia = db.Column(db.String(255), nullable=True) 
    
    veiculo = db.relationship('Veiculo', backref='registros', lazy=True)
    usuario = db.relationship('Usuario', backref='viagens', lazy=True)

# --- FORMULÁRIOS ---
class CadastroUsuarioForm(FlaskForm):
    nome = StringField('Nome Completo', validators=[DataRequired()])
    cpf = StringField('CPF', validators=[DataRequired(), Length(min=11, max=11)])
    senha = PasswordField('Senha', validators=[DataRequired()])
    cargo = SelectField('Cargo', choices=[('Motorista', 'Motorista'), ('Admin', 'Admin')], validators=[DataRequired()])
    gabinete = SelectField('Gabinete Fixo', choices=LISTA_GABINETES, validators=[DataRequired()])
    submit = SubmitField('Cadastrar Usuário')

# app.py (trecho de definição do form)
class RegistroSaidaForm(FlaskForm):
    veiculo_modelo = SelectField('Veículo', choices=[], coerce=int, validators=[DataRequired()])
    km_saida = FloatField('KM Saída', validators=[DataRequired()])
    destino_finalidade = StringField('Destino', validators=[DataRequired()])
    foto_km_saida = FileField('Foto Painel', validators=[FileRequired(), FileAllowed(['jpg','png','jpeg'])])
    submit = SubmitField('Registrar Saída')

class RegistroChegadaForm(FlaskForm):
    registro_id = SelectField('Viagem em Aberto', coerce=int, validators=[DataRequired()])
    km_chegada = FloatField('KM Chegada', validators=[DataRequired()]) 
    foto_km_chegada = FileField('Foto Painel', validators=[FileRequired()])
    submit = SubmitField('Registrar Chegada')

class VeiculoForm(FlaskForm):
    modelo = StringField('Modelo do Veículo', validators=[DataRequired()])
    placa = StringField('Placa', validators=[DataRequired(), Length(min=7, max=8)])
    submit = SubmitField('Salvar Veículo')

class FotoOcorrencia(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    registro_id = db.Column(db.Integer, db.ForeignKey('registro_uso.id'), nullable=False)
    caminho_foto = db.Column(db.String(255), nullable=False)
    
    # Relacionamento para facilitar a busca
    registro = db.relationship('RegistroUso', backref='fotos_ocorrencia_multiplas', lazy=True)



# --- FUNÇÕES AUXILIARES ---
def salvar_foto_compacta(foto_campo, prefixo):
    try:
        if not foto_campo:
            return None
        filename = secure_filename(foto_campo.filename)
        nome_final = f"{prefixo}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
        caminho_completo = os.path.join(app.config['UPLOAD_FOLDER'], nome_final)
        # Garante que o diretório existe
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        img = Image.open(foto_campo)
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        img.thumbnail((1024, 1024))
        img.save(caminho_completo, optimize=True, quality=70)
        print(f"[DEBUG] imagem salva: {caminho_completo}")
        return nome_final
    except Exception as e:
        print(f"[ERRO salvar_foto_compacta]: {e}")
        import traceback; traceback.print_exc()
        return None

# --- ROTAS PRINCIPAIS ---
@app.route('/')
def index():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    viagem_em_aberto = RegistroUso.query.filter_by(usuario_id=current_user.id, km_chegada=None).first()
    return render_template('index.html', tem_viagem_aberta=(viagem_em_aberto is not None))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        cpf = request.form.get('cpf', '').replace('.', '').replace('-', '').strip()
        senha = request.form.get('senha')
        usuario = Usuario.query.filter_by(cpf=cpf).first()
        if usuario and check_password_hash(usuario.senha, senha):
            if not usuario.ativo:
                flash('Conta inativada.', 'danger')
                return redirect(url_for('login'))
            login_user(usuario)
            return redirect(url_for('index'))
        else:
            flash('CPF ou Senha incorretos!', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# --- GESTÃO DE USUÁRIOS ---
@app.route('/gestao-usuarios', methods=['GET', 'POST'])
@login_required
def gestao_usuarios():
    if current_user.cargo != 'Admin':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('index'))
    form = CadastroUsuarioForm()
    if form.validate_on_submit():
        cpf_limpo = form.cpf.data.replace('.', '').replace('-', '').strip()
        existente = Usuario.query.filter_by(cpf=cpf_limpo).first()
        if existente:
            flash('CPF já cadastrado!', 'danger')
        else:
            novo_u = Usuario(
                nome=form.nome.data, cpf=cpf_limpo,
                senha=generate_password_hash(form.senha.data),
                cargo=form.cargo.data, gabinete=form.gabinete.data, ativo=True
            )
            db.session.add(novo_u); db.session.commit()
            flash('Usuário cadastrado!', 'success')
            return redirect(url_for('gestao_usuarios'))
    usuarios = Usuario.query.all()
    return render_template('gestao_usuarios.html', form=form, usuarios=usuarios)

@app.route('/editar-usuario/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_usuario(id):
    if current_user.cargo != 'Admin': return redirect(url_for('index'))
    usuario = db.session.get(Usuario, id)
    if request.method == 'POST':
        usuario.nome = request.form.get('nome')
        usuario.cpf = request.form.get('cpf').replace('.', '').replace('-', '').strip()
        usuario.gabinete = request.form.get('gabinete')
        nova_senha = request.form.get('senha')
        if nova_senha:
            usuario.senha = generate_password_hash(nova_senha)
        db.session.commit()
        flash('Usuário atualizado!', 'success')
        return redirect(url_for('gestao_usuarios'))
    return render_template('editar_usuario.html', usuario=usuario, gabinetes=LISTA_GABINETES)

@app.route('/alternar-status/<int:id>')
@login_required
def alternar_status(id):
    if current_user.cargo != 'Admin': return redirect(url_for('index'))
    usuario = db.session.get(Usuario, id)
    if usuario and usuario.id != current_user.id:
        usuario.ativo = not usuario.ativo
        db.session.commit()
    return redirect(url_for('gestao_usuarios'))

@app.route('/excluir-usuario/<int:id>')
@login_required
def excluir_usuario(id):
    if current_user.cargo != 'Admin': return redirect(url_for('index'))
    usuario = db.session.get(Usuario, id)
    if usuario and usuario.id != current_user.id:
        db.session.delete(usuario)
        db.session.commit()
    return redirect(url_for('gestao_usuarios'))

# --- GESTÃO DE VEÍCULOS ---
@app.route('/gestao-veiculos', methods=['GET', 'POST'])
@login_required
def gestao_veiculos():
    if current_user.cargo != 'Admin': return redirect(url_for('index'))
    form = VeiculoForm()
    if form.validate_on_submit():
        novo_v = Veiculo(modelo=form.modelo.data, placa=form.placa.data.upper())
        db.session.add(novo_v); db.session.commit()
        flash('Veículo cadastrado!', 'success')
        return redirect(url_for('gestao_veiculos'))
    veiculos = Veiculo.query.all()
    return render_template('gestao_veiculos.html', form=form, veiculos=veiculos)

@app.route('/editar_veiculo/<int:id>', methods=['POST'])
@login_required
def editar_veiculo(id):
    if current_user.cargo != 'Admin': return redirect(url_for('index'))
    veiculo = db.session.get(Veiculo, id)
    if veiculo:
        veiculo.modelo = request.form.get('modelo')
        veiculo.placa = request.form.get('placa').upper()
        db.session.commit()
        flash('Veículo atualizado!', 'success')
    return redirect(url_for('gestao_veiculos'))

@app.route('/excluir-veiculo/<int:id>')
@login_required
def excluir_veiculo(id):
    if current_user.cargo != 'Admin': return redirect(url_for('index'))
    v = db.session.get(Veiculo, id)
    if v:
        if v.registros:
            flash('Não é possível excluir veículo com histórico!', 'danger')
        else:
            db.session.delete(v); db.session.commit()
            flash('Veículo removido!', 'warning')
    return redirect(url_for('gestao_veiculos'))

# --- REGISTRO DE VIAGENS ---
# app.py (trecho da rota registrar_saida)
@app.route('/registrar-saida', methods=['GET', 'POST'])
@login_required
def registrar_saida():
    form = RegistroSaidaForm()
    veiculos = Veiculo.query.all()
    
    # Força o banco a ler os dados mais recentes antes de listar ocupados
    db.session.expire_all()
    ocupados = {r.veiculo_id for r in RegistroUso.query.filter_by(km_chegada=None).all()}

    form.veiculo_modelo.choices = [
        (v.id, f"{v.modelo} {'(EM USO)' if v.id in ocupados else ''}") for v in veiculos
    ]

    if form.validate_on_submit():
        selecionado_id = form.veiculo_modelo.data
        if selecionado_id in ocupados:
            flash('Veículo selecionado já está em uso.', 'danger')
            return redirect(url_for('registrar_saida'))

        v = db.session.get(Veiculo, selecionado_id)
        if not v:
            flash('Veículo inválido.', 'danger')
            return redirect(url_for('registrar_saida'))

        novo = RegistroUso(
            usuario_id=current_user.id,
            gabinete_vereador=current_user.gabinete if current_user.gabinete else "Administrativo/Geral",
            motorista_nome=current_user.nome,
            veiculo_id=v.id,
            km_saida=form.km_saida.data,
            foto_km_saida=salvar_foto_compacta(form.foto_km_saida.data, "S"),
            destino_finalidade=form.destino_finalidade.data,
            data_hora_saida=datetime.now()
        )
        
        try:
            db.session.add(novo)
            db.session.commit()
            # AJUSTE 1: Limpa o cache da sessão para a próxima requisição ver o dado novo
            db.session.expire_all() 
            flash('Saída registrada!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao salvar: {e}', 'danger')
            return redirect(url_for('registrar_saida'))

        # AJUSTE 2: Redireciona com um parâmetro para evitar cache do navegador
        return redirect(url_for('index', v=datetime.now().timestamp()))

    return render_template('registrar_saida.html', form=form, veiculos=veiculos, ocupados=list(ocupados))

@app.route('/registrar-chegada', methods=['GET', 'POST'])
@login_required
def registrar_chegada():
    form = RegistroChegadaForm()

    # Busca registros em aberto (km_chegada is None)
    registros_abertos = RegistroUso.query.filter_by(km_chegada=None).order_by(RegistroUso.data_hora_saida.desc()).all()

    # Popula choices do select (valor = id, label descritivo)
    form.registro_id.choices = [
        (r.id, f"{r.veiculo.modelo} - {r.motorista_nome} - {r.data_hora_saida.strftime('%d/%m %H:%M')} - KM {r.km_saida}")
        for r in registros_abertos
    ]

    # Envia lista para o template (o template verifica 'registros')
    registros = registros_abertos

    if form.validate_on_submit():
        reg = db.session.get(RegistroUso, form.registro_id.data)
        if reg:
            # Salva os dados básicos
            reg.km_chegada = form.km_chegada.data
            reg.data_hora_chegada = datetime.now()
            reg.foto_km_chegada = salvar_foto_compacta(form.foto_km_chegada.data, "C")
            reg.observacoes = request.form.get('observacoes')

            # Salva múltiplas fotos de ocorrência (se houver)
            arquivos = request.files.getlist('foto_ocorrencia')
            for f in arquivos:
                if f and f.filename != '':
                    nome_arquivo = salvar_foto_compacta(f, "O")
                    if nome_arquivo:
                        nova_foto = FotoOcorrencia(registro_id=reg.id, caminho_foto=nome_arquivo)
                        db.session.add(nova_foto)
                    else:
                        print(f"[WARN] falha ao salvar foto de ocorrência: {getattr(f, 'filename', '<sem-nome>')}")

            # Atualiza km_atual do veículo se necessário
            try:
                if reg.km_chegada is not None:
                    v = reg.veiculo
                    if v and (v.km_atual is None or reg.km_chegada > v.km_atual):
                        v.km_atual = int(reg.km_chegada)
            except Exception:
                pass

            db.session.commit()
            flash('Chegada registrada!', 'success')
            return redirect(url_for('index'))

    return render_template('registrar_chegada.html', form=form, registros=registros)
        
# --- HISTÓRICO E RELATÓRIOS ---
@app.route('/historico')
@login_required
def historico():
    if current_user.cargo != 'Admin': 
        return redirect(url_for('index'))
    
    # 1. Captura os filtros da URL
    f_inicio = request.args.get('data_inicio')
    f_fim = request.args.get('data_fim')
    f_gabinete = request.args.get('gabinete')
    f_veiculo = request.args.get('veiculo')

    query = RegistroUso.query

    # 2. Aplica filtros de data (SQL)
    if f_inicio:
        dt_ini = datetime.strptime(f_inicio, '%Y-%m-%d')
        query = query.filter(RegistroUso.data_hora_saida >= dt_ini)
    if f_fim:
        dt_fim = datetime.strptime(f_fim, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        query = query.filter(RegistroUso.data_hora_saida <= dt_fim)
    
    if f_gabinete: 
        query = query.filter(RegistroUso.gabinete_vereador == f_gabinete)
    if f_veiculo:
        query = query.filter(RegistroUso.veiculo_id == int(f_veiculo))

    # 3. Executa a busca
    registros = query.order_by(RegistroUso.data_hora_saida.desc()).all()
    
    # 4. CRIAÇÃO DA VARIÁVEL DADOS_GRAFICO (O que estava faltando)
    dados_grafico = {}
    for reg in registros:
        if reg.km_chegada:
            distancia = reg.km_chegada - reg.km_saida
            # Soma a distância para o modelo do veículo
            nome_veiculo = reg.veiculo.modelo
            dados_grafico[nome_veiculo] = dados_grafico.get(nome_veiculo, 0) + distancia

    # 5. Listas auxiliares para os filtros (Selects)
    gabinetes_list = [g[0] for g in LISTA_GABINETES]
    veiculos_list = Veiculo.query.all()

    # 6. Renderiza a página enviando todas as variáveis necessárias
    return render_template('historico.html', 
                           registros=registros, 
                           gabinetes=gabinetes_list, 
                           veiculos=veiculos_list,
                           labels_carros=list(dados_grafico.keys()), 
                           valores_km=list(dados_grafico.values()))

@app.route('/relatorio-ocorrencias')
@login_required
def relatorio_ocorrencias():
    if current_user.cargo != 'Admin': 
        return redirect(url_for('index'))
    
    # Buscamos todas as viagens finalizadas
    viagens = RegistroUso.query.filter(RegistroUso.km_chegada != None).order_by(RegistroUso.data_hora_chegada.desc()).all()
    
    return render_template('relatorio_ocorrencias.html', ocorrencias=viagens)


@app.route('/exportar-excel')
@login_required
def exportar_excel():
    if current_user.cargo != 'Admin': return redirect(url_for('index'))
    
    # --- CAPTURA DOS NOVOS FILTROS DE PERÍODO ---
    f_inicio = request.args.get('data_inicio')
    f_fim = request.args.get('data_fim')
    f_gabinete = request.args.get('gabinete')
    f_veiculo = request.args.get('veiculo')

    query = RegistroUso.query

    # Filtro de Período (SQL)
    if f_inicio:
        dt_ini = datetime.strptime(f_inicio, '%Y-%m-%d')
        query = query.filter(RegistroUso.data_hora_saida >= dt_ini)
    if f_fim:
        dt_fim = datetime.strptime(f_fim, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        query = query.filter(RegistroUso.data_hora_saida <= dt_fim)
    
    if f_gabinete: 
        query = query.filter(RegistroUso.gabinete_vereador == f_gabinete)
    if f_veiculo:
        query = query.filter(RegistroUso.veiculo_id == int(f_veiculo))
    
    registros_filtrados = query.order_by(RegistroUso.data_hora_saida.desc()).all()

    # --- MONTAGEM DOS DADOS (Com Data de Chegada) ---
    dados = []
    for r in registros_filtrados:
        distancia = (r.km_chegada - r.km_saida) if r.km_chegada else 0
        dados.append({
            "DATA/HORA SAÍDA": r.data_hora_saida.strftime('%d/%m/%Y %H:%M'),
            "DATA/HORA CHEGADA": r.data_hora_chegada.strftime('%d/%m/%Y %H:%M') if r.data_hora_chegada else "EM TRÂNSITO",
            "MOTORISTA": r.motorista_nome.upper(),
            "VEÍCULO": f"{r.veiculo.modelo} ({r.veiculo.placa})",
            "GABINETE": r.gabinete_vereador,
            "KM INICIAL": r.km_saida,
            "KM FINAL": r.km_chegada if r.km_chegada else "---",
            "TOTAL KM": distancia,
            "DESTINO/FINALIDADE": r.destino_finalidade
        })

    df = pd.DataFrame(dados)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Relatório')
        workbook = writer.book
        worksheet = writer.sheets['Relatório']

        # Estilos Institucionais
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        signature_line = Border(top=Side(style='medium')) 
        
        # Formatação de Cabeçalho e Ajuste Automático de Colunas
        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                cell.border = thin_border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                val_len = len(str(cell.value))
                if val_len > max_length:
                    max_length = val_len
            worksheet.column_dimensions[column_letter].width = max_length + 5

        # --- LINHAS DE ASSINATURA CORRIGIDAS ---
        last_row = len(registros_filtrados) + 4
        
        # 1. Responsável (Lado Esquerdo)
        worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=3)
        for c in range(1, 4):
            worksheet.cell(row=last_row, column=c).border = signature_line
        
        label_ass = worksheet.cell(row=last_row, column=1)
        label_ass.value = "Assinatura do Responsável (Transportes)"
        label_ass.alignment = Alignment(horizontal='center')

        # 2. Administração (Lado Direito)
        worksheet.merge_cells(start_row=last_row, start_column=7, end_row=last_row, end_column=9)
        for c in range(7, 10):
            worksheet.cell(row=last_row, column=c).border = signature_line
            
        label_adm = worksheet.cell(row=last_row, column=7)
        label_adm.value = "Visto da Administração"
        label_adm.alignment = Alignment(horizontal='center')

        # Rodapé de Emissão
        emissao_cell = worksheet.cell(row=last_row + 2, column=1)
        emissao_cell.value = f"Relatório extraído em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        emissao_cell.font = Font(italic=True, size=9, color='777777')

    output.seek(0)
    nome_doc = f"Relatorio_Frota_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
    return send_file(output, download_name=nome_doc, as_attachment=True)

@app.route('/editar_viagem/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_viagem(id):
    if current_user.cargo != 'Admin': return redirect(url_for('index'))
    viagem = db.session.get(RegistroUso, id)
    if request.method == 'POST':
        viagem.km_saida = float(request.form.get('km_saida'))
        viagem.km_chegada = float(request.form.get('km_chegada')) if request.form.get('km_chegada') else None
        viagem.destino_finalidade = request.form.get('destino')
        db.session.commit(); flash('Viagem atualizada!', 'success')
        return redirect(url_for('historico'))
    return render_template('editar_viagem.html', viagem=viagem)

@app.route('/admin/dashboard')
@login_required
def painel_admin():
    if current_user.cargo != 'Admin':
        flash('Acesso negado!', 'danger')
        return redirect(url_for('index'))

    veiculos = Veiculo.query.all()
    viagens_concluidas = RegistroUso.query.filter(RegistroUso.km_chegada != None).all()

    # 1. Lógica do Ranking de Gabinetes (KM Total)
    ranking = {}
    for v in viagens_concluidas:
        distancia = v.km_chegada - v.km_saida
        ranking[v.gabinete_vereador] = ranking.get(v.gabinete_vereador, 0) + distancia
    
    # Ordenar ranking do maior para o menor
    ranking_ordenado = dict(sorted(ranking.items(), key=lambda item: item[1], reverse=True))

    # 2. Lógica de Atualização de KM Atual dos Veículos
    # (Garante que o KM do veículo seja sempre o maior KM de chegada registrado para ele)
    for v in veiculos:
        ultimo_registro = RegistroUso.query.filter_by(veiculo_id=v.id).filter(RegistroUso.km_chegada != None).order_by(RegistroUso.km_chegada.desc()).first()
        if ultimo_registro:
            v.km_atual = ultimo_registro.km_chegada
    db.session.commit()

    return render_template('painel_admin.html', 
                           veiculos=veiculos, 
                           ranking=ranking_ordenado)

@app.route('/admin/resetar-revisao/<int:id>')
@login_required
def resetar_revisao(id):
    if current_user.cargo != 'Admin': return redirect(url_for('index'))
    v = db.session.get(Veiculo, id)
    if v:
        # Define a próxima revisão para daqui a 10.000 KM a partir do KM atual
        v.km_revisao_proxima = v.km_atual + 10000
        db.session.commit()
        flash(f'Revisão do {v.modelo} atualizada para {v.km_revisao_proxima} KM!', 'success')
    return redirect(url_for('painel_admin'))

@app.route('/download/<path:filename>')
@login_required
def download_file(filename):
    """
    Força o download de um arquivo armazenado no diretório de uploads,
    protegendo contra path traversal.
    """
    upload_folder = current_app.config.get('UPLOAD_FOLDER')
    if not upload_folder:
        abort(404)

    # Caminho absoluto do upload folder e do arquivo solicitado
    upload_folder_abs = os.path.abspath(upload_folder)
    requested_path = os.path.abspath(os.path.join(upload_folder_abs, filename))

    # Protege contra path traversal: requested_path deve ficar dentro de upload_folder_abs
    try:
        common = os.path.commonpath([upload_folder_abs, requested_path])
    except ValueError:
        # caminhos em drives diferentes (Windows) -> bloqueia
        abort(404)

    if common != upload_folder_abs or not os.path.exists(requested_path):
        abort(404)

    # Nome seguro para o cabeçalho de download (mantém o nome original, sanitizado)
    download_name = secure_filename(os.path.basename(requested_path)) or os.path.basename(requested_path)

    # Envia o arquivo forçando download. Compatibilidade com diferentes versões do Flask:
    try:
        return send_file(requested_path, as_attachment=True, download_name=download_name)
    except TypeError:
        return send_file(requested_path, as_attachment=True, attachment_filename=download_name)

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# --- INICIALIZAÇÃO ---
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)

